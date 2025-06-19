import json
from datetime import date
from openpyxl import Workbook
from decimal import Decimal, InvalidOperation, DivisionByZero

from .models import TaxasVendas
from vendas.models import Vendas
from compras.models import Compras
from funcionarios.models import Pagamento

from django.views import View
from django.utils import timezone
from django.shortcuts import render
from django.http import HttpResponse
from django.urls import reverse_lazy
from django.core.paginator import Paginator
from django.core.serializers.json import DjangoJSONEncoder
from django.views.generic import ListView, CreateView, UpdateView
from django.db.models import Sum, Value, CharField, DecimalField, Q, F
from django.db.models.functions import Coalesce, ExtractWeek, ExtractYear
from django.contrib.auth.mixins import LoginRequiredMixin, PermissionRequiredMixin

from project.utils import generate_success_url
from project.mixins import TenantQuerysetMixin, HandleNoPermissionMixin


def exportar_vendas(request):
    # Verifica se o usuário está autenticado e tem tenant
    if not request.user.is_authenticated or not hasattr(request.user, 'tenant'):
        return HttpResponse("Acesso não autorizado", status=403)
    
    tenant = request.user.tenant
    # Filtra as vendas pelo tenant
    queryset = Vendas.objects.filter(tenant=tenant)

    wb = Workbook()
    ws = wb.active
    ws.title = "Vendas"

    # Adiciona cabeçalhos
    headers = ["Data Vendas", "Período", 
                "Rodízio", 
                "Dinheiro", "Pix", 
                "Débito MasterCard", "Débito Visa", "Debito Elo",
                "Crédito MasterCard", "Crédito Visa", "Crédito Elo",
                "Alelo", "American Express", "Hiper", "Sodexo", "Ticket Rest", "Vale Refeição", "DinersClub",
                "Sócio"]
    ws.append(headers)

    for obj in queryset:
        ws.append([obj.data_venda, obj.periodo, 
                    obj.rodizio,
                    obj.dinheiro, obj.pix, 
                    obj.debito_mastercard, obj.debito_visa, obj.debito_elo,
                    obj.credito_mastercard, obj.credito_visa, obj.credito_elo,
                    obj.alelo, obj.american_express, obj.hiper, obj.sodexo, obj.ticket_rest, obj.vale_refeicao, obj.dinersclub,
                    obj.socio])

    # Define o range onde os filtros serão aplicados (A1:M1 no exemplo)
    ws.auto_filter.ref = "A1:S1"  # Ajuste conforme o número de colunas
    # Congelar cabeçalho e linha
    ws.freeze_panes = "C2"

    # Ajustar largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Cria a resposta HTTP com o arquivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=Vendas_{tenant}.xlsx"
    wb.save(response)
    return response

def exportar_compras(request):
    # Verifica se o usuário está autenticado e tem tenant
    if not request.user.is_authenticated or not hasattr(request.user, 'tenant'):
        return HttpResponse("Acesso não autorizado", status=403)
    
    tenant = request.user.tenant
    # Filtra as vendas pelo tenant
    queryset = Compras.objects.filter(tenant=tenant)

    wb = Workbook()
    ws = wb.active
    ws.title = "Compras"

    # Adiciona cabeçalhos
    headers = ["Data Compra", "Data Vencimento", "Data Pagamento",
               "Fornecedor",
               "Valor da Compra", "Valor de Pagamento", "Quantidade", 
               "Boleto",
               "Grupo do Produto", "Produto", "Classificação",
               "Forma de Pagamento",
               "Observação"]
    ws.append(headers)

    # fornecedor é uma chave da tabela Fornecedor
    for obj in queryset:
        # Acessa diretamente o nome_empresa do fornecedor relacionado
        fornecedor_nome = obj.fornecedor.nome_empresa if obj.fornecedor else ""

    for obj in queryset:
        ws.append([obj.data_compra, obj.data_vencimento, obj.data_pagamento, 
                    fornecedor_nome,
                    obj.valor_compra, obj.valor_pago, obj.qtd, 
                    obj.numero_boleto,
                    obj.grupo_produto, obj.produto, obj.classificacao,
                    obj.forma_pagamento,
                    obj.observacao])

    # Define o range onde os filtros serão aplicados (A1:M1 no exemplo)
    ws.auto_filter.ref = "A1:M1"  # Ajuste conforme o número de colunas
    # Congelar cabeçalho
    ws.freeze_panes = "A2"
    # Ajustar largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Cria a resposta HTTP com o arquivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=Compras_{tenant}.xlsx"
    wb.save(response)

    return response

def exportar_pg_funcionarios(request):
    # Verifica se o usuário está autenticado e tem tenant
    if not request.user.is_authenticated or not hasattr(request.user, 'tenant'):
        return HttpResponse("Acesso não autorizado", status=403)
    
    tenant = request.user.tenant
    # Filtra as vendas pelo tenant
    queryset = Pagamento.objects.filter(tenant=tenant)

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagamento de Funcionários"

    # Adiciona cabeçalhos
    headers = ["Data Pagamento", "Nome Funcionário", "Valor Pago", "Tipo de Pagamento", "Forma de Pagamento"]
    ws.append(headers)

    # Adicionando filtro no cabeçalho
    ws.auto_filter.ref = "A1:E1"

    for obj in queryset:
        # Acessa o nome do funcionário através da ForeignKey
        funcionario_nome = obj.nome_funcionario.nome_funcionario if obj.nome_funcionario else ""

    for obj in queryset:
        ws.append([obj.data_pagamento,
                   funcionario_nome,
                   obj.valor_pago, 
                    obj.tipo_pagamento,
                    obj.forma_pagamento])

    # Congelar cabeçalhos
    ws.freeze_panes = "A2"

    # Ajustar largura das colunas automaticamente
    for col in ws.columns:
        max_length = 0
        column = col[0].column_letter  # Get the column name
        
        for cell in col:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2)
        ws.column_dimensions[column].width = adjusted_width

    # Cria a resposta HTTP com o arquivo Excel
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response["Content-Disposition"] = f"attachment; filename=Pagamentos_{tenant}.xlsx"
    wb.save(response)
    return response


class FiltrosFinanceiro:
    def __init__(self, data_inicio=None, data_fim=None, campo_data='data_venda', 
                 periodo=None, 
                 fornecedor=None, classificacao=None,
                 funcionario = None, tipo_pagamento = None, forma_pagamento = None,
                 tenant=None,  **kwargs):
        # Geral
        self.data_inicio = data_inicio
        self.data_fim = data_fim
        self.campo_data = campo_data  # Define qual campo de data será filtrado (padrão é 'data_compra')
        self.tenant = tenant

        # Vendas
        self.periodo = periodo
        
        # Compras
        self.fornecedor = fornecedor
        self.classificacao = classificacao
        
        # Funcionarios
        self.funcionario = funcionario
        self.tipo_pagamento = tipo_pagamento
        self.forma_pagamento = forma_pagamento

        self.kwargs = kwargs  # Outros filtros adicionais passados como kwargs

    def aplicar_filtros(self, queryset):
        # Filtrar pelo tenant
        if self.tenant:
            queryset = queryset.filter(tenant=self.tenant)

        if self.data_inicio:
            queryset = queryset.filter(**{f"{self.campo_data}__gte": self.data_inicio})
        
        if self.data_fim:
            queryset = queryset.filter(**{f"{self.campo_data}__lte": self.data_fim})

        if self.periodo:  # Aplicando o filtro de período
            queryset = queryset.filter(periodo__in=self.periodo)

        if self.fornecedor:
            # Remove valores vazios da lista de fornecedores
            fornecedores_validos = [f for f in self.fornecedor if f]
            if fornecedores_validos:
                queryset = queryset.filter(fornecedor__id__in=fornecedores_validos)

        if self.classificacao:
            # Remove valores vazios da lista de classificações
            classificacoes_validas = [c for c in self.classificacao if c]
            if classificacoes_validas:
                queryset = queryset.filter(classificacao__in=classificacoes_validas)
        
        if self.funcionario:
            funcionarios_validos = [f for f in self.funcionario if f]
            if funcionarios_validos:
                queryset = queryset.filter(nome_funcionario__id__in=funcionarios_validos)

        if self.tipo_pagamento:
            tipo_pagamentos_validos = [t for t in self.tipo_pagamento if t]
            if tipo_pagamentos_validos:
                queryset = queryset.filter(tipo_pagamento__in=tipo_pagamentos_validos)

        if self.forma_pagamento:
            forma_pagamentos_validos = [f for f in self.forma_pagamento if f]
            if forma_pagamentos_validos:
                queryset = queryset.filter(forma_pagamento__in=forma_pagamentos_validos)
        
        # Adicionando outros filtros passados como kwargs
        for key, value in self.kwargs.items():
            if value:
                queryset = queryset.filter(**{key: value})

        return queryset


class DashboardBaseView(View):
    def calcular_totais_pagamentos(self, pagamentos_queryset):
        total_pagamentos = pagamentos_queryset.aggregate(total_despesas=Sum('valor_pago'))['total_despesas'] or 0
        return total_pagamentos
    
    def calcular_totais_vendas(self, vendas_queryset):
        vendas_total = vendas_queryset.aggregate(
        total_rodizio=Coalesce(Sum('rodizio', output_field=DecimalField()), 0, output_field=DecimalField()),
        total_dinheiro=Coalesce(Sum('dinheiro', output_field=DecimalField()), 0, output_field=DecimalField()),
        total_pix=Coalesce(Sum('pix', output_field=DecimalField()), 0, output_field=DecimalField()),
        total_debito=Coalesce(Sum('debito_mastercard', output_field=DecimalField()), 0, output_field=DecimalField()) +
                     Coalesce(Sum('debito_visa', output_field=DecimalField()), 0, output_field=DecimalField()) +
                     Coalesce(Sum('debito_elo', output_field=DecimalField()), 0, output_field=DecimalField()),
        total_credito=Coalesce(Sum('credito_mastercard', output_field=DecimalField()), 0, output_field=DecimalField()) +
                      Coalesce(Sum('credito_visa', output_field=DecimalField()), 0, output_field=DecimalField()) +
                      Coalesce(Sum('credito_elo', output_field=DecimalField()), 0, output_field=DecimalField()),
        total_beneficio=Coalesce(Sum('alelo', output_field=DecimalField()), 0, output_field=DecimalField()) +
                        Coalesce(Sum('american_express', output_field=DecimalField()), 0, output_field=DecimalField()) +
                        Coalesce(Sum('hiper', output_field=DecimalField()), 0, output_field=DecimalField()) +
                        Coalesce(Sum('sodexo', output_field=DecimalField()), 0, output_field=DecimalField()) +
                        Coalesce(Sum('ticket_rest', output_field=DecimalField()), 0, output_field=DecimalField()) +
                        Coalesce(Sum('vale_refeicao', output_field=DecimalField()), 0, output_field=DecimalField()) +
                        Coalesce(Sum('dinersclub', output_field=DecimalField()), 0, output_field=DecimalField()),
        total_socio=Coalesce(Sum('socio', output_field=DecimalField()), 0, output_field=DecimalField())
    )

        # Use valores padrão (0) para None
        total_dinheiro = vendas_total['total_dinheiro']
        total_pix = vendas_total['total_pix']
        total_debito = vendas_total['total_debito']
        total_credito = vendas_total['total_credito']
        total_beneficio = vendas_total['total_beneficio']
        total_socio = vendas_total['total_socio']
        total_rodizio = vendas_total['total_rodizio']

        # Cálculo total de vendas
        total_vendas = total_dinheiro + total_pix + total_debito + total_credito + total_beneficio

        # Calcule as taxas
        taxa_dinheiro = (total_dinheiro / total_vendas * 100) if total_vendas != 0 else 0
        taxa_pix = (total_pix / total_vendas * 100) if total_vendas != 0 else 0
        taxa_debito = (total_debito / total_vendas * 100) if total_vendas != 0 else 0
        taxa_credito = (total_credito / total_vendas * 100) if total_vendas != 0 else 0
        taxa_beneficio = (total_beneficio / total_vendas * 100) if total_vendas != 0 else 0

        # Ticket médio
        ticket_medio = (total_vendas / total_rodizio) if total_rodizio != 0 else 0

        return {
            'total_dinheiro': total_dinheiro,
            'total_pix': total_pix,
            'total_debito': total_debito,
            'total_credito': total_credito,
            'total_beneficio': total_beneficio,
            'total_socio': total_socio,
            'total_vendas': total_vendas,
            'total_rodizio': total_rodizio,
            'ticket_medio': ticket_medio,
            'taxa_dinheiro': taxa_dinheiro,
            'taxa_pix': taxa_pix,
            'taxa_debito': taxa_debito,
            'taxa_credito': taxa_credito,
            'taxa_beneficio': taxa_beneficio,
        }

    def calcular_taxas_vendas(self, vendas_queryset):
        # Define as taxas por bandeira de cartão
        # taxas_bandeiras = {
        #     "debito_mastercard": 0.0095,
        #     "debito_visa": 0.0095,
        #     "debito_elo": 0.0098,
        #     "credito_mastercard": 0.0299,
        #     "credito_visa": 0.0299,
        #     "credito_elo": 0.0299,
        #     "hiper": 0.0299,
        #     "dinersclub": 0.0299,
        #     "american_express": 0.025,
        #     "alelo": 0.07,
        #     "sodexo": 0.069,
        #     "vale_refeicao": 0.069,
        #     "ticket_rest": 0.07,
        # }


        # Configura precisão decimal
        # getcontext().prec = 8

        # Verifica se existem vendas
        if not vendas_queryset.exists():
            return Decimal('0.00')

        # Obtém todas as taxas em ordem cronológica (como lista)
        taxas = list(TaxasVendas.objects.order_by('data_taxa_venda'))
        
        if not taxas:
            return Decimal('0.00')

        total_taxas = Decimal('0.00')
        
        # Processa cada taxa
        for i in range(len(taxas)):
            taxa = taxas[i]
            data_inicio = taxa.data_taxa_venda
            data_fim = taxas[i+1].data_taxa_venda if i+1 < len(taxas) else None
            
            # Cria nova queryset para o período
            vendas_filtradas = vendas_queryset.model.objects.filter(
                pk__in=vendas_queryset.values_list('pk', flat=True),
                data_venda__gte=data_inicio
            )
            
            if data_fim:
                vendas_filtradas = vendas_filtradas.filter(data_venda__lt=data_fim)
            
            # Calcula para cada bandeira
            for bandeira in [
                'debito_mastercard', 'debito_visa', 'debito_elo',
                'credito_mastercard', 'credito_visa', 'credito_elo',
                'hiper', 'dinersclub', 'american_express',
                'alelo', 'sodexo', 'vale_refeicao', 'ticket_rest'
            ]:
                valor_taxa = getattr(taxa, bandeira, Decimal('0.00'))
                valor_taxa = valor_taxa if valor_taxa is not None else Decimal('0.00')
                
                if valor_taxa != Decimal('0.00'):
                    resultado = vendas_filtradas.aggregate(
                        total=Sum(F(bandeira) * valor_taxa / 100,
                        output_field=DecimalField()
                    ))

                    total_bandeira = Decimal(resultado.get('total') or '0.00')
                    total_taxas += total_bandeira

        return total_taxas

    def calcular_totais_compras(self, compras_queryset):
        total_compras = compras_queryset.aggregate(total_despesas=Sum('valor_compra'))['total_despesas'] or 0
        total_pago = compras_queryset.aggregate(total_despesas=Sum('valor_pago'))['total_despesas'] or 0
        cmv = compras_queryset.filter(classificacao='CMV').aggregate(total_cmv=Sum('valor_compra'))['total_cmv'] or 0
        gasto_fixo = compras_queryset.filter(classificacao='Gasto Fixo').aggregate(total_gasto_fixo=Sum('valor_compra'))['total_gasto_fixo'] or 0
        gasto_variavel = compras_queryset.filter(classificacao='Gasto Variável').aggregate(total_gasto_variavel=Sum('valor_compra'))['total_gasto_variavel'] or 0
        
        # Taxas
        taxa_cmv = round(cmv / total_compras * 100, 2) if total_compras != 0 else 0
        taxa_gasto_fixo = round(gasto_fixo / total_compras * 100, 2) if total_compras != 0 else 0
        taxa_gasto_variavel = round(gasto_variavel / total_compras * 100, 2) if total_compras != 0 else 0

        return {
            'total_compras': total_compras,
            'total_pago': total_pago,
            'total_cmv': cmv,
            'total_gasto_fixo': gasto_fixo,
            'total_gasto_variavel': gasto_variavel,
            'taxa_cmv': taxa_cmv,
            'taxa_gasto_fixo': taxa_gasto_fixo,
            'taxa_gasto_variavel': taxa_gasto_variavel,
        }

    def calcular_contas_vencidas(self, compras_queryset):
        # Filtra as compras com data de vencimento menor que a data atual e onde valor_pago é zero ou não existe
        compras_vencidas_nao_pagas = compras_queryset.filter(
            data_vencimento__lt=date.today(),  # Vencidas
            data_pagamento__isnull=True  # Não pagas (data_pagamento é None)
        )

        # Soma o valor total das compras vencidas e não pagas
        total_compras_vencidas = compras_vencidas_nao_pagas.aggregate(
            total_vencidas_nao_pagas=Sum('valor_compra')
        )['total_vencidas_nao_pagas'] or 0

        return total_compras_vencidas
    
    def calcular_contas_dentro_do_prazo(self, compras_queryset):
        # Filtra as compras com data de vencimento maior ou igual à data atual e onde data_pagamento está vazio
        compras_dentro_do_prazo = compras_queryset.filter(
            data_vencimento__gte=date.today(),  # Dentro do prazo (vencimento maior ou igual a hoje)
            data_pagamento__isnull=True  # Não pagas (data_pagamento é None)
        )

        # Soma o valor total das compras não pagas dentro do prazo
        total_compras_dentro_do_prazo = compras_dentro_do_prazo.aggregate(
            total_nao_pagas_dentro_do_prazo=Sum('valor_compra')
        )['total_nao_pagas_dentro_do_prazo'] or 0

        return total_compras_dentro_do_prazo

    


class TaxaListView(LoginRequiredMixin, PermissionRequiredMixin, TenantQuerysetMixin, HandleNoPermissionMixin, ListView):
    model = TaxasVendas
    template_name = 'financeiro/taxasvendas_list.html'  # Garante que o template correto será usado
    context_object_name = 'taxas_list'  # Isso define o nome que será usado no template
    permission_required = 'financeiro.view_taxasvendas'  # Permissão para visualizar objetos'


class TaxaCreateView(LoginRequiredMixin, PermissionRequiredMixin, TenantQuerysetMixin, HandleNoPermissionMixin, CreateView):
    model = TaxasVendas
    fields = ["data_taxa_venda", "debito_mastercard", "debito_visa", "debito_elo", "credito_mastercard", 
              "credito_visa", "credito_elo", "hiper", "dinersclub", "american_express", 
              "alelo", "sodexo", "vale_refeicao", "ticket_rest"]
    permission_required = 'financeiro.add_taxasvendas'
    
    def get_form(self, form_class=None):
        form = super().get_form(form_class)
        tenant = getattr(self.request.user, 'tenant', None)
        if tenant:
            form.instance.tenant = tenant
            form.instance.author = self.request.user
        return form  # Adicionado o return

    def form_valid(self, form):
        tenant = getattr(self.request.user, 'tenant', None)
        if not tenant:
            return self.form_invalid(form)
        form.instance.tenant = tenant
        form.instance.author = self.request.user
        return super().form_valid(form)
    
    def get_success_url(self):
        return reverse_lazy("taxas_vendas")


class TaxaUpDateView(LoginRequiredMixin, PermissionRequiredMixin, TenantQuerysetMixin, HandleNoPermissionMixin, UpdateView):
    model = TaxasVendas
    fields = ["data_taxa_venda", "debito_mastercard", "debito_visa", "debito_elo", "credito_mastercard", "credito_visa", "credito_elo",
              "hiper", "dinersclub", "american_express", "alelo", "sodexo", "vale_refeicao", "ticket_rest"]
    success_url = reverse_lazy("taxas_vendas")
    permission_required = 'financeiro.change_taxasvendas'  # Permissão para visualizar objetos'

    def form_valid(self, form):
        # Certifique-se de que o tenant está sendo atribuído corretamente, se necessário
        tenant = getattr(self.request.user, 'tenant', None)
        if not tenant:
            return self.form_invalid(form)  # Retorna erro se o tenant não for encontrado
        form.instance.tenant = tenant
        return super().form_valid(form)

    def get_success_url(self):
        tenant = getattr(self.request.user, 'tenant', None)
        return generate_success_url('taxas_vendas', tenant)


class DashboardVendasView(DashboardBaseView):
    def get(self, request):
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        periodo = request.GET.get('periodo')
        tenant = getattr(request.user, 'tenant', None)  # Obter o tenant do usuário logado

        # Verificar e converter o período para lista
        if periodo:
            periodo = periodo.split(',')

        vendas = Vendas.objects.filter(tenant=self.request.user.tenant)
        filtros = FiltrosFinanceiro(
            data_inicio=data_inicio, 
            data_fim=data_fim, 
            periodo=periodo,
            tenant=tenant
        )
        
        # Aplicar filtros no queryset de vendas
        vendas_filtradas = filtros.aplicar_filtros(vendas)
        # Ordenar do maior para o menor (datas decrescentes)
        vendas_filtradas = vendas_filtradas.order_by('-data_venda')

        # Paginação
        paginator = Paginator(vendas_filtradas, 31)  # Exibe 10 registros por página
        page_number = self.request.GET.get('page')
        vendas_paginadas = paginator.get_page(page_number)

        totais_vendas = self.calcular_totais_vendas(vendas_paginadas.object_list)
        total_taxas = self.calcular_taxas_vendas(vendas_paginadas.object_list) or 0
        # Calcular a taxa percentual de total_taxa em relação ao total_vendas
        taxa_total_taxa = (total_taxas / totais_vendas['total_vendas'] * 100) if totais_vendas['total_vendas'] != 0 else 0

        total_compras = self.calcular_totais_compras(Compras.objects.filter(tenant=self.request.user.tenant))
        total_pagamento_funcionarios = self.calcular_totais_pagamentos(Pagamento.objects.filter(tenant=self.request.user.tenant))

        context = {
            **totais_vendas,
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'campo_data': 'data_venda',
            'periodo': periodo,
            'total_taxas': total_taxas,
            'taxa_total_taxa': taxa_total_taxa,
            'total_despesas': total_compras,
            'pagamento_funcionario': total_pagamento_funcionarios,
            'vendas_por_forma': vendas_filtradas,
            'vendas_paginadas': vendas_paginadas,
        }
        return render(request, 'financeiro/dashboard_vendas.html', context)


class DashboardComprasView(DashboardBaseView):
    def get(self, request):
        today = timezone.now().date()
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        classificacao_selecionado = request.GET.getlist('classificacao')
        fornecedores_selecionado = request.GET.getlist('fornecedor')
        tenant = getattr(request.user, 'tenant', None)  # Obter o tenant do usuário logado

        # Widget RADIO
        filtrar_vencidas = request.GET.get('filtrar_vencidas') == 'on'

        compras = Compras.objects.filter(tenant=self.request.user.tenant).order_by('-data_compra')
        filtros = FiltrosFinanceiro(
            data_inicio=data_inicio, 
            data_fim=data_fim, 
            campo_data='data_compra',
            classificacao=classificacao_selecionado,
            fornecedor=fornecedores_selecionado,
            tenant=tenant)
        
        # Aplique os filtros na queryset
        compras_filtradas = filtros.aplicar_filtros(compras)

        # Ordenar do maior para o menor (datas decrescentes)
        # compras_filtradas = compras_filtradas.order_by('-data_compra')

        if filtrar_vencidas:
            compras_filtradas = compras_filtradas.filter(
                Q(data_vencimento__lt=today) & Q(data_pagamento__isnull=True)
            )

        # Paginação
        paginator = Paginator(compras_filtradas, 10)  # Exibe 10 registros por página
        page_number = self.request.GET.get('page')
        compras_paginadas = paginator.get_page(page_number)

        totais_compras = self.calcular_totais_compras(compras_filtradas)
        fornecedores = Compras.objects.filter(tenant=tenant).values('fornecedor__id', 'fornecedor__nome_empresa').distinct()
        classificacao = Compras.objects.filter(tenant=tenant).values('classificacao').distinct()

        context = {
            **totais_compras,
            'today': today, # Data atual para comparação com vencimentos
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'campo_data': 'data_compra',
            'compras': compras_filtradas,
            'fornecedores_selecionados': fornecedores_selecionado,
            'fornecedores': fornecedores,
            'classificacao_selecionado': classificacao_selecionado,
            'classificacao': classificacao,
            'filtrar_vencidas': filtrar_vencidas, #Filtro de radio
            'compras_paginadas': compras_paginadas,
        }
        return render(request, 'financeiro/dashboard_compras.html', context)


class DashboardFuncionariosView(DashboardBaseView):
    def get(self, request):
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        funcionario_selecionado = request.GET.getlist('nome_funcionario')
        tipo_pagamento_selecionado = request.GET.getlist('tipo_pagamento')
        forma_pagamento_selecionado = request.GET.getlist('forma_pagamento')
        tenant = getattr(request.user, 'tenant', None)  # Obter o tenant do usuário logado

        funcionarios = Pagamento.objects.filter(tenant=self.request.user.tenant).order_by('-data_pagamento')

        filtros = FiltrosFinanceiro(
            data_inicio=data_inicio,
            data_fim=data_fim,
            campo_data='data_pagamento',
            funcionario=funcionario_selecionado,
            tipo_pagamento=tipo_pagamento_selecionado,
            forma_pagamento=forma_pagamento_selecionado,
            tenant=tenant
            )
        
        # Aplicar filtros no queryset de funcionários
        funcionarios_filtrado = filtros.aplicar_filtros(funcionarios)
        # Ordenar do maior para o menor (datas decrescentes)
        # funcionarios_filtrado = funcionarios_filtrado.order_by('-data_pagamento')

        tipo_pagamentos_agrupados = funcionarios_filtrado.values('tipo_pagamento').annotate(total_valor_pago=Sum('valor_pago'))
        forma_pagamentos_agrupados = funcionarios_filtrado.values('forma_pagamento').annotate(total_valor_pago=Sum('valor_pago'))

        # Paginação
        paginator = Paginator(funcionarios_filtrado, 10)  # Exibe 10 registros por página
        page_number = self.request.GET.get('page')
        funcionarios_paginados = paginator.get_page(page_number)
        
        # Totais de pagamentos
        total_pagamentos = self.calcular_totais_pagamentos(funcionarios_paginados.object_list)
        total_tipo_pagamento = sum(item['total_valor_pago'] for item in tipo_pagamentos_agrupados)
        total_forma_pagamento = sum(item['total_valor_pago'] for item in forma_pagamentos_agrupados)

        # Dados para o filtro
        # nome_funcionario = Pagamento.objects.values('nome_funcionario__id', 'nome_funcionario__nome_funcionario').distinct()
        nome_funcionario = Pagamento.objects.filter(tenant=tenant).values('nome_funcionario__id', 'nome_funcionario__nome_funcionario').distinct()
        tipos_pagamento = Pagamento.objects.filter(tenant=tenant).values('tipo_pagamento').distinct()
        formas_pagamento = Pagamento.objects.filter(tenant=tenant).values('forma_pagamento').distinct()

        # Agrupar pagamentos por tipo de pagamento e funcionário
        funcionarios_por_tipo = (
            funcionarios_filtrado.values('tipo_pagamento', 'nome_funcionario__nome_funcionario')
            .annotate(total_valor_pago=Sum('valor_pago'))
            .order_by('tipo_pagamento', 'nome_funcionario__nome_funcionario')
        )

        # Agrupar pagamentos por forma pagamento e funcionário
        funcionarios_por_forma = (
            funcionarios_filtrado.values('forma_pagamento', 'nome_funcionario__nome_funcionario')
            .annotate(total_valor_pago=Sum('valor_pago'))
            .order_by('forma_pagamento', 'nome_funcionario__nome_funcionario')
        )

        # Agregar valores pagos por cargo
        pagamentos_por_cargo = (
            funcionarios_filtrado.values('nome_funcionario__contratacao__cargo')
            .annotate(total_valor_pago=Sum('valor_pago'))
            .order_by('nome_funcionario__contratacao__cargo')
        )

        # Listar funcionários e seus valores pagos, agrupados por cargo
        funcionarios_por_cargo = (
            funcionarios_filtrado.values(
                'nome_funcionario__contratacao__cargo',
                'nome_funcionario__nome_funcionario',
                'valor_pago'
            )
            .order_by('nome_funcionario__contratacao__cargo')
        )

        # Agregar valores pagos por setor
        pagamentos_por_setor = (
            funcionarios_filtrado.values('nome_funcionario__contratacao__setor')
            .annotate(total_valor_pago=Sum('valor_pago'))
            .order_by('nome_funcionario__contratacao__setor')
        )

        # Listar funcionários e seus valores pagos, agrupados por cargo
        funcionario_por_setor = (
            funcionarios_filtrado.values(
                'nome_funcionario__contratacao__setor',
                'nome_funcionario__nome_funcionario',
                'valor_pago'
            )
            .order_by('nome_funcionario__contratacao__setor')
        )

        context = {
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            'campo_data': 'data_pagamento',
            'funcionarios': funcionarios_filtrado,
            'nome_funcionarios': nome_funcionario,
            'funcionario_selecionados': funcionario_selecionado,
            'tipo_pagamento_selecionado': tipo_pagamento_selecionado,
            'forma_pagamento_selecionado': forma_pagamento_selecionado,
            'total_pagamentos': total_pagamentos,
            'tipos_pagamento': tipos_pagamento,
            'formas_pagamento': formas_pagamento,
            'tipo_pagamentos_agrupados': tipo_pagamentos_agrupados,
            'forma_pagamentos_agrupados': forma_pagamentos_agrupados,
            'total_tipo_pagamento': total_tipo_pagamento,
            'total_forma_pagamento': total_forma_pagamento,
            'funcionarios_por_tipo': funcionarios_por_tipo,
            'funcionarios_por_forma': funcionarios_por_forma,
            'pagamentos_por_cargo': pagamentos_por_cargo,
            'funcionarios_por_cargo': funcionarios_por_cargo,
            'pagamentos_por_setor': pagamentos_por_setor,
            'funcionario_por_setor': funcionario_por_setor,
            'funcionarios_paginados': funcionarios_paginados,
        }
        return render(request, 'financeiro/dashboard_funcionarios.html', context)


class DashboardResumoView(DashboardBaseView):
    def get(self, request):
        # Filtros e dados de vendas
        data_inicio = request.GET.get('data_inicio')
        data_fim = request.GET.get('data_fim')
        tenant = getattr(request.user, 'tenant', None)  # Obter o tenant do usuário logado

        vendas = Vendas.objects.filter(tenant=tenant)
        filtros_vendas = FiltrosFinanceiro(
            data_inicio=data_inicio, 
            data_fim=data_fim
        )
        vendas_filtradas = filtros_vendas.aplicar_filtros(vendas)
        totais_vendas = self.calcular_totais_vendas(vendas_filtradas)

        total_taxas = self.calcular_taxas_vendas(vendas_filtradas) or 0
        fundo_caixa = totais_vendas['total_vendas'] * Decimal('0.06') or Decimal('0')

        # Ticket médio diário para gráfico de linha
        ticket_medio_por_data = {}
        for v in vendas_filtradas:
            if v.data_venda:
                valor_venda = (
                    (v.dinheiro or 0) + (v.pix or 0) + (v.debito_mastercard or 0) + (v.debito_visa or 0) +
                    (v.debito_elo or 0) + (v.credito_mastercard or 0) + (v.credito_visa or 0) +
                    (v.credito_elo or 0) + (v.alelo or 0) + (v.american_express or 0) + (v.hiper or 0) +
                    (v.sodexo or 0) + (v.ticket_rest or 0) + (v.vale_refeicao or 0) + (v.dinersclub or 0) +
                    (v.socio or 0)
                )
                if v.data_venda not in ticket_medio_por_data:
                    ticket_medio_por_data[v.data_venda] = {'total_venda': 0, 'total_rodizio': 0}
                ticket_medio_por_data[v.data_venda]['total_venda'] += float(valor_venda)
                ticket_medio_por_data[v.data_venda]['total_rodizio'] += float(v.rodizio or 0)

        # Calcule a média de vendas por dia usando ticket_medio_por_data
        if ticket_medio_por_data:
            total_vendas_por_dia = [info['total_venda'] for info in ticket_medio_por_data.values()]
            media_vendas = sum(total_vendas_por_dia) / len(total_vendas_por_dia)
        else:
            media_vendas = 0

        ticket_labels = []
        ticket_data = []
        for data_venda in sorted(ticket_medio_por_data.keys()):
            total_venda = ticket_medio_por_data[data_venda]['total_venda']
            total_rodizio = ticket_medio_por_data[data_venda]['total_rodizio']
            ticket_labels.append(data_venda.strftime('%d/%m/%Y'))
            ticket_data.append(round(total_venda / total_rodizio, 2) if total_rodizio else 0)

        # Média de rodízio por dia
        if ticket_medio_por_data:
            rodizio_por_dia = [info['total_rodizio'] for info in ticket_medio_por_data.values()]
            media_rodizio = sum(rodizio_por_dia) / len(rodizio_por_dia)
        else:
            media_rodizio = 0

        # Gráfico de pizza para almoço e jantar
        total_almoco = sum(
            float(
                (v.dinheiro or 0) + (v.pix or 0) + (v.debito_mastercard or 0) + (v.debito_visa or 0) +
                (v.debito_elo or 0) + (v.credito_mastercard or 0) + (v.credito_visa or 0) +
                (v.credito_elo or 0) + (v.alelo or 0) + (v.american_express or 0) + (v.hiper or 0) +
                (v.sodexo or 0) + (v.ticket_rest or 0) + (v.vale_refeicao or 0) + (v.dinersclub or 0) +
                (v.socio or 0)
            )
            for v in vendas_filtradas if hasattr(v, 'periodo') and v.periodo == 'Almoço'
        )
        total_jantar = sum(
            float(
                (v.dinheiro or 0) + (v.pix or 0) + (v.debito_mastercard or 0) + (v.debito_visa or 0) +
                (v.debito_elo or 0) + (v.credito_mastercard or 0) + (v.credito_visa or 0) +
                (v.credito_elo or 0) + (v.alelo or 0) + (v.american_express or 0) + (v.hiper or 0) +
                (v.sodexo or 0) + (v.ticket_rest or 0) + (v.vale_refeicao or 0) + (v.dinersclub or 0) +
                (v.socio or 0)
            )
            for v in vendas_filtradas if hasattr(v, 'periodo') and v.periodo == 'Jantar'
        )

        donut_data = [total_almoco, total_jantar]

        # Filtros e dados de compras
        classificacao_selecionada = request.GET.getlist('classificacao_compras')
        fornecedores_selecionados = request.GET.getlist('fornecedor_compras')

        compras = Compras.objects.filter(tenant=tenant)
        filtros_compras = FiltrosFinanceiro(
            data_inicio=data_inicio, 
            data_fim=data_fim, 
            campo_data='data_compra',
            classificacao=classificacao_selecionada,
            fornecedor=fornecedores_selecionados,
            tenant=tenant,
        )
        compras_filtradas = filtros_compras.aplicar_filtros(compras)
        totais_compras = self.calcular_totais_compras(compras_filtradas)
        total_compras_vencidas = self.calcular_contas_vencidas(compras_filtradas)
        # Calcular o total das contas não pagas dentro do prazo
        total_compras_dentro_do_prazo = self.calcular_contas_dentro_do_prazo(compras_filtradas)

        # Filtra apenas compras do grupo_produto='Peixe'
        compras_peixe = compras_filtradas.filter(grupo_produto='Peixe')

        # Agrupa por produto e soma o valor_compra
        segmentos_peixe = (
            compras_peixe
            .values('produto')
            .annotate(valor=Sum('valor_compra'))
            .order_by('-valor')
        )

        # Prepara listas para o gráfico
        segmento_labels = [item['produto'] for item in segmentos_peixe]
        segmento_data = [float(item['valor'] or 0) for item in segmentos_peixe]

        # Filtros e dados de pagamentos de funcionários
        data_inicio_funcionarios = request.GET.get('data_inicio_funcionarios')
        data_fim_funcionarios = request.GET.get('data_fim_funcionarios')
        funcionario_selecionado = request.GET.getlist('nome_funcionario')
        tipo_pagamento_selecionado = request.GET.getlist('tipo_pagamento')
        forma_pagamento_selecionado = request.GET.getlist('forma_pagamento')

        funcionarios = Pagamento.objects.filter(tenant=tenant)
        filtros_funcionarios = FiltrosFinanceiro(
            data_inicio=data_inicio,
            data_fim=data_fim,
            campo_data='data_pagamento',
            funcionario=funcionario_selecionado,
            tipo_pagamento=tipo_pagamento_selecionado,
            forma_pagamento=forma_pagamento_selecionado,
        )

        funcionarios_filtrados = filtros_funcionarios.aplicar_filtros(funcionarios)
        total_pagamentos_funcionarios = self.calcular_totais_pagamentos(funcionarios_filtrados)
        
        taxa_pagamentos_funcionarios = total_pagamentos_funcionarios / totais_vendas['total_vendas'] * 100 if totais_vendas['total_vendas'] != 0 else 0
        taxa_pagamentos_funcionarios = round(taxa_pagamentos_funcionarios, 2)

        total_rescisao = (
        funcionarios_filtrados.filter(tipo_pagamento='Rescisão')
        .aggregate(total=Sum('valor_pago'))['total'] or 0
    )

        taxa_rescisao = total_rescisao / totais_vendas['total_vendas'] * 100 if totais_vendas['total_vendas'] != 0 else 0
        taxa_rescisao = round(taxa_rescisao, 2)

        # Agregar dados de compras por classificação, tipo de produto e produto
        gastos_por_classificacao = compras_filtradas.values(
            'classificacao').annotate(
            total_valor=Sum('valor_compra')).order_by(
            'classificacao')

        gastos_por_grupo_produto = compras_filtradas.values(
            'classificacao', 'grupo_produto').annotate(
            total_valor=Sum('valor_compra')).order_by(
            'classificacao', 'grupo_produto')

        gastos_por_produto = compras_filtradas.values(
            'classificacao', 'grupo_produto', 'produto').annotate(
            total_valor=Sum('valor_compra')).order_by(
            'classificacao', 'grupo_produto', 'produto')

        # Agregar dados de funcionários (considerando "Mão de Obra" como classificação)
        gastos_funcionarios_por_classificacao = funcionarios_filtrados.annotate(
            classificacao=Value("Mão de Obra", output_field=CharField())
        ).values(
            'classificacao').annotate(
            total_valor=Sum('valor_pago')).order_by(
            'classificacao')

        gastos_funcionarios_por_grupo = funcionarios_filtrados.annotate(
            classificacao=Value("Mão de Obra", output_field=CharField()),
            grupo_produto=F('tipo_pagamento')
        ).values(
            'classificacao', 'grupo_produto').annotate(
            total_valor=Sum('valor_pago')).order_by(
            'classificacao', 'grupo_produto')

        gastos_funcionarios_por_produto = funcionarios_filtrados.annotate(
            classificacao=Value("Mão de Obra", output_field=CharField()),
            grupo_produto=F('tipo_pagamento'),
            produto=F('nome_funcionario__nome_funcionario')  # foi usado __ devido nome_funcionario ser um chave estrangeira 
        ).values(
            'classificacao', 'grupo_produto', 'produto').annotate(
            total_valor=Sum('valor_pago')).order_by(
            'classificacao', 'grupo_produto', 'produto')

        # Combinar os resultados de compras e funcionários
        gastos_por_classificacao = list(gastos_por_classificacao) + list(gastos_funcionarios_por_classificacao)
        gastos_por_grupo_produto = list(gastos_por_grupo_produto) + list(gastos_funcionarios_por_grupo)
        gastos_por_produto = list(gastos_por_produto) + list(gastos_funcionarios_por_produto)

        # Calcular totais combinados (compras + funcionários)
        # total_compras = sum(item['total_valor'] for item in gastos_por_classificacao if item['classificacao'] != "Mão de Obra")
        # total_funcionarios = sum(item['total_valor'] for item in gastos_por_classificacao if item['classificacao'] == "Mão de Obra") or 0


        total_compras = sum(item['total_valor'] or 0 for item in gastos_por_classificacao if item['classificacao'] != "Mão de Obra")
        total_funcionarios = sum(item['total_valor'] or 0 for item in gastos_por_classificacao if item['classificacao'] == "Mão de Obra")



        total_geral = total_compras + total_funcionarios

        # Capturando o parâmetro da radio button
        usar_totais_vendas = request.GET.get('usar_totais_vendas') == 'false'

        # Calcular taxas percentuais para tabela gastos
        # for classificacao in gastos_por_classificacao:
        #     if usar_totais_vendas:
        #         classificacao['taxa_percentual'] = (classificacao['total_valor'] / total_geral * 100) if total_geral != 0 else 0
        #     else:
        #         classificacao['taxa_percentual'] = (classificacao['total_valor'] / totais_vendas['total_vendas'] * 100) if totais_vendas['total_vendas'] != 0 else 0

        # for grupo_produto in gastos_por_grupo_produto:
        #     for classificacao in gastos_por_classificacao:
        #         if grupo_produto['classificacao'] == classificacao['classificacao']:
        #             if usar_totais_vendas:
        #                 grupo_produto['taxa_percentual'] = (grupo_produto['total_valor'] / classificacao['total_valor'] * 100) if classificacao['total_valor'] != 0 else 0
        #             else:
        #                 grupo_produto['taxa_percentual'] = (grupo_produto['total_valor'] / totais_vendas['total_vendas'] * 100) if totais_vendas['total_vendas'] != 0 else 0
        #             break

        # for produto in gastos_por_produto:
        #     for grupo_produto in gastos_por_grupo_produto:
        #         if (produto['classificacao'] == grupo_produto['classificacao'] and 
        #             produto['grupo_produto'] == grupo_produto['grupo_produto']):
        #             if usar_totais_vendas:
        #                 produto['taxa_percentual'] = (produto['total_valor'] / grupo_produto['total_valor'] * 100) if grupo_produto['total_valor'] != 0 else 0
        #             else:
        #                 produto['taxa_percentual'] = (produto['total_valor'] / totais_vendas['total_vendas'] * 100) if totais_vendas['total_vendas'] != 0 else 0
        #             break

        def calcular_percentual(numerador, denominador):
            """Calcula percentual de forma segura pois quando o valor no banco era None a página dava erro"""
            try:
                # Converte para Decimal se não for None
                num = Decimal(numerador) if numerador is not None else Decimal('0')
                den = Decimal(denominador) if denominador is not None else Decimal('0')
                
                # Evita divisão por zero
                if den == 0:
                    return Decimal('0')
                    
                return (num / den) * 100
            except (TypeError, InvalidOperation, DivisionByZero):
                return Decimal('0')

        # Processamento das classificações
        for classificacao in gastos_por_classificacao:
            denominador = total_geral if usar_totais_vendas else totais_vendas.get('total_vendas', 0)
            classificacao['taxa_percentual'] = calcular_percentual(
                classificacao.get('total_valor'), 
                denominador
            )

        # Processamento dos grupos de produto
        for grupo_produto in gastos_por_grupo_produto:
            for classificacao in gastos_por_classificacao:
                if grupo_produto['classificacao'] == classificacao['classificacao']:
                    denominador = (
                        classificacao.get('total_valor') 
                        if usar_totais_vendas 
                        else totais_vendas.get('total_vendas', 0)
                    )
                    grupo_produto['taxa_percentual'] = calcular_percentual(
                        grupo_produto.get('total_valor'),
                        denominador
                    )
                    break

        # Processamento dos produtos
        for produto in gastos_por_produto:
            for grupo_produto in gastos_por_grupo_produto:
                if (produto['classificacao'] == grupo_produto['classificacao'] and 
                    produto['grupo_produto'] == grupo_produto['grupo_produto']):
                    denominador = (
                        grupo_produto.get('total_valor') 
                        if usar_totais_vendas 
                        else totais_vendas.get('total_vendas', 0)
                    )
                    produto['taxa_percentual'] = calcular_percentual(
                        produto.get('total_valor'),
                        denominador
                    )
                    break

        # # Dados para o gráfico de barras - Compras por grupo de produto
        # bar_grupo_labels = [item['grupo_produto'] for item in gastos_por_grupo_produto]
        # bar_grupo_data = [float(item['total_valor']) for item in gastos_por_grupo_produto]

        # # Dados para o gráfico de teia (Radar) - Compras por classificação
        # radar_labels = [item['classificacao'] for item in gastos_por_classificacao]
        # radar_data = [float(item['total_valor']) for item in gastos_por_classificacao]

        # Função auxiliar para conversão segura desta forma não da erro nos gráficos quando o valor do banco for None
        def safe_float(value, default=0.0):
            """Converte para float de forma segura, tratando None e outros casos"""
            try:
                return float(value) if value is not None else default
            except (TypeError, ValueError):
                return default

        # Dados para o gráfico de barras - Compras por grupo de produto
        bar_grupo_labels = [item['grupo_produto'] for item in gastos_por_grupo_produto]
        bar_grupo_data = [safe_float(item.get('total_valor')) for item in gastos_por_grupo_produto]

        # Dados para o gráfico de teia (Radar) - Compras por classificação
        radar_labels = [item['classificacao'] for item in gastos_por_classificacao]
        radar_data = [safe_float(item.get('total_valor')) for item in gastos_por_classificacao]



        # Dados para o gráfico de linha - Compras por data
        compras_por_data = {}
        for c in compras_filtradas:
            if c.data_compra:
                compras_por_data.setdefault(c.data_compra, 0)
                compras_por_data[c.data_compra] += float(c.valor_compra or 0)

        linha_labels = [data.strftime('%d/%m/%Y') for data in sorted(compras_por_data.keys())]
        linha_data = [compras_por_data[data] for data in sorted(compras_por_data.keys())]

        # Filtre apenas CMV
        cmv_queryset = compras_filtradas.filter(classificacao='CMV')

        # Agrupe por semana e ano
        cmv_semanal = (
            cmv_queryset
            .annotate(ano=ExtractYear('data_compra'), semana=ExtractWeek('data_compra'))
            .values('ano', 'semana')
            .annotate(total=Sum('valor_compra'))
            .order_by('ano', 'semana')
        )

        # Prepare os labels e dados para o gráfico de barras
        cmv_labels = [f"Sem {item['semana']}/{item['ano']}" for item in cmv_semanal]
        cmv_data = [float(item['total']) for item in cmv_semanal]

        cmv_por_data = {}
        for c in compras_filtradas:
            if c.classificacao == 'CMV' and c.data_compra:
                cmv_por_data.setdefault(c.data_compra, 0)
                cmv_por_data[c.data_compra] += float(c.valor_compra or 0)

        cmv_bar_labels = [data.strftime('%d/%m/%Y') for data in sorted(cmv_por_data.keys())]
        cmv_bar_data = [cmv_por_data[data] for data in sorted(cmv_por_data.keys())]

        # Cálculo do lucro líquido
        lucro_liquido = totais_vendas['total_vendas'] - totais_compras['total_compras'] - total_pagamentos_funcionarios
        taxa_lucro = lucro_liquido / totais_vendas['total_vendas'] * 100 if totais_vendas['total_vendas'] != 0 else 0
        taxa_lucro = round(taxa_lucro, 2)

        datas_unicas = sorted({v.data_venda for v in vendas_filtradas if v.data_venda})
        periodos = ['Almoço', 'Jantar']

        # Inicializar estrutura: {data: {periodo: valor}}
        vendas_por_data_periodo = {data: {p: 0 for p in periodos} for data in datas_unicas}

        for v in vendas_filtradas:
            if v.data_venda and v.periodo in periodos:
                valor_venda = (
                    (v.dinheiro or 0) + (v.pix or 0) + (v.debito_mastercard or 0) + (v.debito_visa or 0) +
                    (v.debito_elo or 0) + (v.credito_mastercard or 0) + (v.credito_visa or 0) +
                    (v.credito_elo or 0) + (v.alelo or 0) + (v.american_express or 0) + (v.hiper or 0) +
                    (v.sodexo or 0) + (v.ticket_rest or 0) + (v.vale_refeicao or 0) + (v.dinersclub or 0) +
                    (v.socio or 0)
                )
                vendas_por_data_periodo[v.data_venda][v.periodo] += float(valor_venda)

        stacked_labels = [data.strftime('%d/%m/%Y') for data in datas_unicas]
        almoco_data = [vendas_por_data_periodo[data]['Almoço'] for data in datas_unicas]
        jantar_data = [vendas_por_data_periodo[data]['Jantar'] for data in datas_unicas]

        # Contexto para o template
        context = {
            **totais_compras,
            **totais_vendas,
            # Dados de filtro
            'data_inicio': data_inicio,
            'data_fim': data_fim,
            # Dados de vendas
            'totais_vendas': totais_vendas['total_vendas'],
            'media_vendas': media_vendas,
            'total_taxas': total_taxas,
            'lucro_liquido': lucro_liquido,
            'taxa_lucro': taxa_lucro,
            'fundo_caixa': fundo_caixa,

            # Dados de compras
            'totais_compras': totais_compras['total_compras'],
            'total_pago_compras': totais_compras['total_pago'],
            'cmv': totais_compras['total_cmv'],
            'gasto_fixo': totais_compras['total_gasto_fixo'],
            'gasto_variavel': totais_compras['total_gasto_variavel'],
            'total_compras_vencidas': total_compras_vencidas,
            'total_compras_dentro_do_prazo': total_compras_dentro_do_prazo,
            'gastos_por_classificacao': gastos_por_classificacao,
            'gastos_por_grupo_produto': gastos_por_grupo_produto,
            'gastos_por_produto': gastos_por_produto,
            'segmento_labels_json': json.dumps(segmento_labels, cls=DjangoJSONEncoder),
            'segmento_data_json': json.dumps(segmento_data, cls=DjangoJSONEncoder),

            # Dados de pagamentos de funcionários
            'total_pagamentos_funcionarios': total_pagamentos_funcionarios,
            'data_inicio_funcionarios': data_inicio_funcionarios,
            'data_fim_funcionarios': data_fim_funcionarios,
            'funcionario_selecionado': funcionario_selecionado,
            'tipo_pagamento_selecionado': tipo_pagamento_selecionado,
            'forma_pagamento_selecionado': forma_pagamento_selecionado,
            'total_rescisao': total_rescisao,
            'taxa_rescisao': taxa_rescisao,
            'taxa_pagamentos_funcionarios': taxa_pagamentos_funcionarios,

            # Lucro líquido
            'lucro_liquido': lucro_liquido,

            # Dados do gráfico
            'stacked_labels_json': json.dumps(stacked_labels, cls=DjangoJSONEncoder),
            'almoco_data_json': json.dumps(almoco_data, cls=DjangoJSONEncoder),
            'jantar_data_json': json.dumps(jantar_data, cls=DjangoJSONEncoder),

            'donut_data': json.dumps([total_almoco, total_jantar], cls=DjangoJSONEncoder),
            
            'ticket_labels_json': json.dumps(ticket_labels, cls=DjangoJSONEncoder),
            'ticket_data_json': json.dumps(ticket_data, cls=DjangoJSONEncoder),

            'radar_labels_json': json.dumps(radar_labels, cls=DjangoJSONEncoder),
            'radar_data_json': json.dumps(radar_data, cls=DjangoJSONEncoder),
            'linha_labels_json': json.dumps(linha_labels, cls=DjangoJSONEncoder),
            'linha_data_json': json.dumps(linha_data, cls=DjangoJSONEncoder),

            'cmv_labels_json': json.dumps(cmv_labels, cls=DjangoJSONEncoder),
            'cmv_data_json': json.dumps(cmv_data, cls=DjangoJSONEncoder),

            'cmv_bar_labels_json': json.dumps(cmv_bar_labels, cls=DjangoJSONEncoder),
            'cmv_bar_data_json': json.dumps(cmv_bar_data, cls=DjangoJSONEncoder),

            'media_rodizio': media_rodizio,

            'bar_grupo_labels_json': json.dumps(bar_grupo_labels, cls=DjangoJSONEncoder),
            'bar_grupo_data_json': json.dumps(bar_grupo_data, cls=DjangoJSONEncoder),
        }

        return render(request, 'financeiro/dashboard_resumo.html', context)